# -*- coding: utf-8 -*-
"""
excel.py
========

Camada de acesso à planilha Excel.

Responsabilidades:

* abrir o ``.xlsx`` preservando 100% da formatação original;
* localizar dinamicamente a linha de cabeçalho e mapear as colunas;
* criar as colunas **Confiança** e **Observação** quando ausentes, herdando o
  estilo do cabeçalho existente;
* ler as empresas, identificando quais já possuem telefone válido;
* gravar resultados **sem jamais apagar conteúdo existente**;
* colorir a linha conforme o desfecho da pesquisa;
* salvar em disco de forma atômica após cada empresa;
* gerar o relatório das empresas sem contato.

Todas as operações de escrita são serializadas por um ``threading.RLock``,
pois o motor de pesquisa trabalha com múltiplas threads.
"""

from __future__ import annotations

import logging
import os
import shutil
import threading
from copy import copy
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
import utils
from modelos import Confianca, Empresa, ResultadoPesquisa, StatusPesquisa

logger = logging.getLogger("localizador.excel")


class PlanilhaError(utils.LocalizadorError):
    """Erro relacionado à leitura ou escrita da planilha."""


class PlanilhaContatos:
    """
    Encapsula a planilha de clientes.

    A planilha **de origem nunca é modificada**: o arquivo é copiado para
    ``Empresas_Preenchidas.xlsx`` na pasta de saída e todas as gravações
    acontecem sobre essa cópia.

    Example:
        >>> planilha = PlanilhaContatos("carteira.xlsx", Path("saida"))
        >>> planilha.abrir()
        >>> empresas = planilha.empresas_pendentes()
    """

    def __init__(self, caminho_origem: str | Path, pasta_saida: Path) -> None:
        self.caminho_origem = Path(caminho_origem)
        self.pasta_saida = Path(pasta_saida)
        self.caminho_saida = self.pasta_saida / config.ARQ_SAIDA_PLANILHA

        self._wb: Optional[Workbook] = None
        self._ws: Optional[Worksheet] = None
        self._lock = threading.RLock()

        self.linha_cabecalho: int = 1
        self.colunas: Dict[str, int] = {}
        self._gravacoes_pendentes = 0

    # ==================================================================
    # Abertura e preparação
    # ==================================================================

    def abrir(self) -> None:
        """
        Copia a planilha para a pasta de saída, abre a cópia e prepara colunas.

        Se ``Empresas_Preenchidas.xlsx`` já existir, ela é reaproveitada — o
        que permite retomar um processamento interrompido sem perder o que já
        havia sido preenchido.

        Raises:
            PlanilhaError: arquivo inexistente, corrompido ou sem as colunas
                mínimas (Razão Social e Contato).
        """
        with self._lock:
            if not self.caminho_origem.is_file():
                raise PlanilhaError(f"Planilha não encontrada: {self.caminho_origem}")

            self.pasta_saida.mkdir(parents=True, exist_ok=True)

            if self.caminho_saida.exists():
                logger.info(
                    "Reaproveitando planilha de saída existente: %s", self.caminho_saida
                )
            else:
                shutil.copy2(self.caminho_origem, self.caminho_saida)
                logger.info("Planilha copiada para %s", self.caminho_saida)

            try:
                # keep_vba mantém macros em .xlsm; data_only=False preserva fórmulas.
                self._wb = load_workbook(
                    self.caminho_saida,
                    data_only=False,
                    keep_vba=self.caminho_saida.suffix.lower() == ".xlsm",
                )
            except Exception as exc:
                raise PlanilhaError(f"Não foi possível abrir a planilha: {exc}") from exc

            self._ws = self._escolher_aba(self._wb)
            self.linha_cabecalho, self.colunas = self._mapear_colunas(self._ws)

            faltando = [c for c in ("razao_social", "contato") if c not in self.colunas]
            if faltando:
                raise PlanilhaError(
                    "A planilha não possui as colunas obrigatórias: "
                    + ", ".join(COLUNA_ROTULO[c] for c in faltando)
                )

            self._garantir_colunas_extras()
            self.salvar(forcar=True)

    def _escolher_aba(self, wb: Workbook) -> Worksheet:
        """
        Seleciona a aba de trabalho.

        Prefere a aba cujo cabeçalho contenha "razão social"; se nenhuma
        atender, usa a aba ativa.
        """
        for aba in wb.worksheets:
            _, colunas = self._mapear_colunas(aba)
            if "razao_social" in colunas and "contato" in colunas:
                logger.info("Aba selecionada: %r", aba.title)
                return aba
        logger.warning("Nenhuma aba com cabeçalho reconhecido; usando a aba ativa.")
        return wb.active

    def _mapear_colunas(self, ws: Worksheet) -> Tuple[int, Dict[str, int]]:
        """
        Localiza a linha de cabeçalho e mapeia chave lógica -> índice de coluna.

        Percorre as 10 primeiras linhas procurando aquela com mais rótulos
        reconhecidos — tolera planilhas com título/logo acima da tabela.
        """
        melhor_linha, melhor_mapa = 1, {}

        for linha in range(1, min(11, ws.max_row + 1)):
            mapa: Dict[str, int] = {}
            for coluna in range(1, min(ws.max_column, 60) + 1):
                valor = ws.cell(row=linha, column=coluna).value
                if valor is None:
                    continue
                chave = _identificar_coluna(str(valor))
                if chave and chave not in mapa:
                    mapa[chave] = coluna
            if len(mapa) > len(melhor_mapa):
                melhor_linha, melhor_mapa = linha, mapa

        return melhor_linha, melhor_mapa

    def _garantir_colunas_extras(self) -> None:
        """Cria as colunas Confiança e Observação, se ainda não existirem."""
        ws = self._exigir_ws()
        modelo = ws.cell(row=self.linha_cabecalho, column=self.colunas["razao_social"])

        for chave in config.COLUNAS_CRIADAS_AUTOMATICAMENTE:
            if chave in self.colunas:
                continue
            nova = max(self.colunas.values()) + 1 if self.colunas else 1
            # Garante que não sobrescreve nada já preenchido.
            while ws.cell(row=self.linha_cabecalho, column=nova).value not in (None, ""):
                nova += 1

            celula = ws.cell(row=self.linha_cabecalho, column=nova)
            celula.value = COLUNA_ROTULO[chave]
            _copiar_estilo(modelo, celula)
            ws.column_dimensions[get_column_letter(nova)].width = (
                14 if chave == "confianca" else 46
            )
            self.colunas[chave] = nova
            logger.info(
                "Coluna %r criada em %s.", COLUNA_ROTULO[chave], get_column_letter(nova)
            )

    # ==================================================================
    # Leitura
    # ==================================================================

    def empresas(self) -> List[Empresa]:
        """
        Lê todas as empresas da planilha.

        Linhas cuja razão social esteja vazia são ignoradas.
        """
        with self._lock:
            ws = self._exigir_ws()
            lista: List[Empresa] = []

            for linha in range(self.linha_cabecalho + 1, ws.max_row + 1):
                razao = self._ler(linha, "razao_social")
                if not razao:
                    continue
                lista.append(
                    Empresa(
                        linha=linha,
                        razao_social=utils.limpar_espacos(razao),
                        cidade=utils.limpar_espacos(self._ler(linha, "cidade")),
                        regiao=utils.limpar_espacos(self._ler(linha, "regiao")),
                        contato_existente=self._ler(linha, "contato"),
                        uf=config.UF_PADRAO,
                    )
                )

            logger.info("%d empresas lidas da planilha.", len(lista))
            return lista

    def empresas_pendentes(self) -> List[Empresa]:
        """
        Empresas que ainda precisam de pesquisa.

        Conforme o escopo, empresas cuja coluna "Contato" já contenha um
        telefone válido são ignoradas. Linhas com apenas e-mail continuam
        pendentes (falta o telefone).
        """
        return [e for e in self.empresas() if not possui_telefone_valido(e.contato_existente)]

    def _ler(self, linha: int, chave: str) -> str:
        """Lê uma célula pela chave lógica da coluna (string vazia se ausente)."""
        coluna = self.colunas.get(chave)
        if not coluna:
            return ""
        valor = self._exigir_ws().cell(row=linha, column=coluna).value
        return "" if valor is None else str(valor).strip()

    # ==================================================================
    # Escrita
    # ==================================================================

    def aplicar_resultado(self, resultado: ResultadoPesquisa, salvar: bool = True) -> None:
        """
        Grava o resultado da pesquisa na linha da empresa.

        Regras aplicadas (todas exigidas no escopo):

        * conteúdo existente **nunca** é apagado — novos contatos são anexados
          ao que já estava na célula;
        * apenas dados de confiança Alta ou Média são gravados; confiança Baixa
          gera somente a observação "Necessita conferência manual.";
        * a linha inteira é colorida conforme o desfecho;
        * ``Confiança`` e ``Observação`` são sempre atualizadas.

        Args:
            resultado: Resultado consolidado da pesquisa.
            salvar: Se ``True``, persiste o arquivo imediatamente.
        """
        with self._lock:
            ws = self._exigir_ws()
            linha = resultado.empresa.linha

            if resultado.status not in (StatusPesquisa.IGNORADO, StatusPesquisa.CANCELADO):
                self._gravar_contato(ws, linha, resultado)
                self._gravar_metadados(ws, linha, resultado)

            cor = resultado.status.cor_linha
            if cor:
                self._pintar_linha(ws, linha, cor)

            self._gravacoes_pendentes += 1

        if salvar:
            self.salvar()

    def _gravar_contato(self, ws: Worksheet, linha: int, res: ResultadoPesquisa) -> None:
        """Anexa telefones/e-mails à coluna Contato preservando o conteúdo atual."""
        novo_texto = res.celula_contato()
        if not novo_texto:
            return

        coluna = self.colunas["contato"]
        celula = ws.cell(row=linha, column=coluna)
        atual = "" if celula.value is None else str(celula.value).strip()

        # Deduplica: não repete um contato que já esteja na célula.
        existentes_digitos = {
            t.digitos for t in utils.extrair_telefones(atual)
        }
        existentes_emails = set(utils.extrair_emails(atual))

        linhas_novas: List[str] = []
        for valor in novo_texto.split("\n"):
            valor = valor.strip()
            if not valor:
                continue
            if "@" in valor:
                if valor.lower() in existentes_emails:
                    continue
            else:
                digitos = utils.so_digitos(valor)
                if digitos in existentes_digitos:
                    continue
            linhas_novas.append(valor)

        if not linhas_novas:
            return

        celula.value = "\n".join([atual] + linhas_novas) if atual else "\n".join(linhas_novas)
        celula.alignment = Alignment(
            wrap_text=True,
            vertical=celula.alignment.vertical or "center",
            horizontal=celula.alignment.horizontal,
        )
        # Ajusta a altura da linha para caber os múltiplos números.
        qtd = celula.value.count("\n") + 1
        if qtd > 1:
            ws.row_dimensions[linha].height = max(
                ws.row_dimensions[linha].height or 15, 15 * qtd
            )

    def _gravar_metadados(self, ws: Worksheet, linha: int, res: ResultadoPesquisa) -> None:
        """Preenche as colunas Confiança e Observação."""
        if "confianca" in self.colunas:
            celula = ws.cell(row=linha, column=self.colunas["confianca"])
            celula.value = self._texto_confianca(res)
            celula.alignment = Alignment(horizontal="center", vertical="center")

        if "observacao" in self.colunas:
            celula = ws.cell(row=linha, column=self.colunas["observacao"])
            texto = self._texto_observacao(res)
            atual = "" if celula.value is None else str(celula.value).strip()
            # Nunca apaga observação preexistente escrita por um humano.
            celula.value = f"{atual} | {texto}" if atual and atual != texto else texto
            celula.alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _texto_confianca(res: ResultadoPesquisa) -> str:
        """Rótulo da coluna Confiança para o resultado."""
        if res.status == StatusPesquisa.REVISAO_MANUAL:
            return ""
        if not res.tem_algum_dado:
            return ""
        return res.confianca.value

    @staticmethod
    def _texto_observacao(res: ResultadoPesquisa) -> str:
        """
        Monta a observação da linha.

        Sempre registra a fonte usada — requisito de rastreabilidade do escopo.
        """
        if res.status == StatusPesquisa.REVISAO_MANUAL:
            base = config.TEXTO_REVISAO_MANUAL
            return f"{base}. {res.observacao}".strip() if res.observacao else base

        if res.status == StatusPesquisa.ERRO:
            return f"Erro na pesquisa: {utils.truncar(res.erro, 120)}"

        if not res.tem_algum_dado:
            return config.TEXTO_NADA_ENCONTRADO

        partes: List[str] = []
        if res.confianca == Confianca.BAIXA:
            partes.append(config.TEXTO_CONFERENCIA_MANUAL)

        fonte = res.fonte_principal
        if fonte:
            partes.append(f"Fonte: {fonte.value}")
        if res.url_principal:
            partes.append(f"URL: {utils.truncar(res.url_principal, 120)}")
        if res.site:
            partes.append(f"Site: {res.site.valor}")
        if res.whatsapps:
            partes.append("WhatsApp: " + ", ".join(w.valor for w in res.whatsapps[:2]))
        if res.empresa.cnpj:
            partes.append(f"CNPJ: {utils.formatar_cnpj(res.empresa.cnpj)}")
        if res.observacao:
            partes.append(res.observacao)

        return " | ".join(partes)

    def _pintar_linha(self, ws: Worksheet, linha: int, cor_argb: str) -> None:
        """Aplica preenchimento sólido a todas as colunas mapeadas da linha."""
        preenchimento = PatternFill(
            start_color=cor_argb, end_color=cor_argb, fill_type="solid"
        )
        ultima = max(self.colunas.values()) if self.colunas else ws.max_column
        for coluna in range(1, ultima + 1):
            ws.cell(row=linha, column=coluna).fill = preenchimento

    # ==================================================================
    # Persistência
    # ==================================================================

    def salvar(self, forcar: bool = False) -> None:
        """
        Salva a planilha em disco de forma atômica.

        Escreve primeiro em um arquivo temporário e só então substitui o
        definitivo (``os.replace``). Assim, uma queda de energia no meio da
        gravação não corrompe o resultado acumulado.
        """
        with self._lock:
            if self._wb is None:
                return
            if not forcar and self._gravacoes_pendentes == 0:
                return

            temporario = self.caminho_saida.with_suffix(
                self.caminho_saida.suffix + ".tmp"
            )
            try:
                self._wb.save(temporario)
                os.replace(temporario, self.caminho_saida)
                self._gravacoes_pendentes = 0
            except PermissionError as exc:
                raise PlanilhaError(
                    f"Não foi possível salvar {self.caminho_saida.name}. "
                    "O arquivo está aberto no Excel? Feche-o e tente novamente."
                ) from exc
            except Exception as exc:
                logger.exception("Falha ao salvar a planilha.")
                raise PlanilhaError(f"Falha ao salvar a planilha: {exc}") from exc
            finally:
                if temporario.exists():
                    try:
                        temporario.unlink()
                    except OSError:
                        pass

    def fechar(self) -> None:
        """Salva e libera o arquivo."""
        with self._lock:
            try:
                self.salvar(forcar=True)
            finally:
                if self._wb is not None:
                    try:
                        self._wb.close()
                    except Exception:
                        pass
                    self._wb = None
                    self._ws = None

    # ==================================================================
    # Relatório
    # ==================================================================

    def gerar_relatorio(self, resultados: List[ResultadoPesquisa]) -> Optional[Path]:
        """
        Gera o relatório das empresas sem contato encontrado.

        Produz um ``.xlsx`` e um ``.csv`` equivalentes, contendo as empresas com
        status "Não encontrado", "Revisão Manual Necessária" ou "Erro".

        Returns:
            Caminho do ``.xlsx`` gerado, ou ``None`` se não houve pendências.
        """
        alvo = {
            StatusPesquisa.NAO_ENCONTRADO,
            StatusPesquisa.REVISAO_MANUAL,
            StatusPesquisa.CONFERENCIA_MANUAL,
            StatusPesquisa.ERRO,
        }
        pendentes = [r for r in resultados if r.status in alvo]
        if not pendentes:
            logger.info("Nenhuma empresa pendente — relatório não gerado.")
            return None

        cabecalho = [
            "Linha", "Razão Social", "Cidade", "Região", "Status",
            "Motivo", "Fontes consultadas", "Tempo (s)",
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = "Sem Contato"
        ws.append(cabecalho)
        for celula in ws[1]:
            celula.font = Font(bold=True)
            celula.fill = PatternFill("solid", start_color="FFD9D9D9")

        linhas_csv: List[List[str]] = []
        for res in pendentes:
            motivo = (
                res.erro
                or res.observacao
                or (
                    config.TEXTO_REVISAO_MANUAL
                    if res.status == StatusPesquisa.REVISAO_MANUAL
                    else config.TEXTO_NADA_ENCONTRADO
                )
            )
            linha = [
                res.empresa.linha,
                res.empresa.razao_social,
                res.empresa.cidade,
                res.empresa.regiao,
                res.status.value,
                utils.truncar(motivo, 250),
                utils.truncar(", ".join(res.fontes_consultadas), 250),
                round(res.duracao, 1),
            ]
            ws.append(linha)
            linhas_csv.append([str(v) for v in linha])

        for coluna, largura in zip("ABCDEFGH", (8, 48, 22, 18, 26, 60, 60, 12)):
            ws.column_dimensions[coluna].width = largura
        ws.freeze_panes = "A2"

        destino = self.pasta_saida / config.ARQ_RELATORIO_SEM_CONTATO
        wb.save(destino)
        wb.close()

        # CSV equivalente, útil para importação em CRM.
        import csv

        destino_csv = self.pasta_saida / config.ARQ_RELATORIO_CSV
        with open(destino_csv, "w", newline="", encoding="utf-8-sig") as arq:
            escritor = csv.writer(arq, delimiter=";")
            escritor.writerow(cabecalho)
            escritor.writerows(linhas_csv)

        logger.info("Relatório gerado: %s (%d empresas).", destino, len(pendentes))
        return destino

    # ==================================================================
    # Internos
    # ==================================================================

    def _exigir_ws(self) -> Worksheet:
        if self._ws is None:
            raise PlanilhaError("A planilha não foi aberta. Chame abrir() antes.")
        return self._ws

    def __enter__(self) -> "PlanilhaContatos":
        self.abrir()
        return self

    def __exit__(self, *_exc) -> None:
        self.fechar()


# ===========================================================================
# Funções auxiliares de módulo
# ===========================================================================

# Rótulo canônico usado ao criar cada coluna.
COLUNA_ROTULO: Dict[str, str] = {
    chave: rotulos[0] for chave, rotulos in config.COLUNAS_PLANILHA.items()
}

# Índice reverso: rótulo normalizado -> chave lógica.
_INDICE_ROTULOS: Dict[str, str] = {
    utils.normalizar(rotulo): chave
    for chave, rotulos in config.COLUNAS_PLANILHA.items()
    for rotulo in rotulos
}


def _identificar_coluna(texto_cabecalho: str) -> Optional[str]:
    """Traduz o texto de um cabeçalho para a chave lógica da coluna."""
    return _INDICE_ROTULOS.get(utils.normalizar(texto_cabecalho))


def _copiar_estilo(origem, destino) -> None:
    """Copia o estilo visual de uma célula para outra (openpyxl exige cópia)."""
    if origem.has_style:
        destino.font = copy(origem.font)
        destino.border = copy(origem.border)
        destino.fill = copy(origem.fill)
        destino.number_format = origem.number_format
        destino.protection = copy(origem.protection)
        destino.alignment = copy(origem.alignment)


def inspecionar(caminho: str | Path) -> Tuple[List[Empresa], int]:
    """
    Lê a planilha em modo somente leitura, sem criar nada em disco.

    Usada pela interface para mostrar o resumo ("X empresas, Y pendentes")
    logo após a seleção do arquivo — nesse momento ainda não faz sentido gerar
    ``Empresas_Preenchidas.xlsx``.

    Args:
        caminho: Caminho da planilha de origem.

    Returns:
        Tupla ``(empresas, quantidade_com_telefone)``.

    Raises:
        PlanilhaError: arquivo inexistente, ilegível ou sem as colunas mínimas.
    """
    origem = Path(caminho)
    if not origem.is_file():
        raise PlanilhaError(f"Planilha não encontrada: {origem}")

    try:
        wb = load_workbook(origem, data_only=True, read_only=True)
    except Exception as exc:
        raise PlanilhaError(f"Não foi possível abrir a planilha: {exc}") from exc

    try:
        for aba in wb.worksheets:
            linhas = list(aba.iter_rows(values_only=True))
            if not linhas:
                continue

            # Procura o cabeçalho nas 10 primeiras linhas.
            indice_cabecalho, colunas = -1, {}
            for indice, linha in enumerate(linhas[:10]):
                mapa = {}
                for posicao, valor in enumerate(linha):
                    if valor is None:
                        continue
                    chave = _identificar_coluna(str(valor))
                    if chave and chave not in mapa:
                        mapa[chave] = posicao
                if len(mapa) > len(colunas):
                    indice_cabecalho, colunas = indice, mapa

            if "razao_social" not in colunas or "contato" not in colunas:
                continue

            def _celula(linha, chave: str) -> str:
                posicao = colunas.get(chave)
                if posicao is None or posicao >= len(linha) or linha[posicao] is None:
                    return ""
                return str(linha[posicao]).strip()

            empresas, com_telefone = [], 0
            for deslocamento, linha in enumerate(linhas[indice_cabecalho + 1:], start=1):
                razao = _celula(linha, "razao_social")
                if not razao:
                    continue
                contato = _celula(linha, "contato")
                if possui_telefone_valido(contato):
                    com_telefone += 1
                empresas.append(
                    Empresa(
                        linha=indice_cabecalho + 1 + deslocamento,
                        razao_social=utils.limpar_espacos(razao),
                        cidade=utils.limpar_espacos(_celula(linha, "cidade")),
                        regiao=utils.limpar_espacos(_celula(linha, "regiao")),
                        contato_existente=contato,
                    )
                )
            return empresas, com_telefone

        raise PlanilhaError(
            "Nenhuma aba com as colunas obrigatórias (Razão Social e Contato)."
        )
    finally:
        wb.close()


def possui_telefone_valido(texto: str) -> bool:
    """
    Diz se o texto de uma célula de contato já contém um telefone válido.

    Um e-mail sozinho **não** conta como telefone: a empresa segue pendente.
    """
    return bool(utils.extrair_telefones(texto or ""))
